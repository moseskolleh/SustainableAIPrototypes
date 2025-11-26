#!/usr/bin/env python3
"""
Generate Feature Table from Partner Feedback
Creates both Excel and Markdown tables with all features
"""

import pandas as pd
from datetime import datetime

# Define all features extracted from partner feedback
features = [
    # Magic Mirror - Department Display Features
    {
        "Feature": "Department-level Metrics Display",
        "Short Description": "Display aggregated AI usage metrics at department level instead of personal data",
        "Prototype": "Magic Mirror (Prototype 1)",
        "Notes": "Removes privacy concerns by showing collective data only. Compliance-friendly approach.",
        "Suggested by": "Matthijs, Jop, Thomas"
    },
    {
        "Feature": "Prompt Efficiency Metrics",
        "Short Description": "Show average prompts per query and efficiency scores for departments",
        "Prototype": "Magic Mirror (Prototype 1)",
        "Notes": "Key metric for measuring sustainable prompt usage. Target: <1.5 prompts per query",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "Energy Usage Tracking",
        "Short Description": "Display real-time energy consumption from AI operations",
        "Prototype": "Magic Mirror (Prototype 1)",
        "Notes": "Makes abstract sustainability concrete through measurable energy data",
        "Suggested by": "Thomas, Matthijs"
    },
    {
        "Feature": "Tool Diversity Metrics",
        "Short Description": "Track variety of AI tools being used across departments",
        "Prototype": "Magic Mirror (Prototype 1)",
        "Notes": "Encourages exploration of more efficient alternatives",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "CO₂ Impact Visualization",
        "Short Description": "Show carbon footprint of AI usage with real-world equivalents",
        "Prototype": "Magic Mirror (Prototype 1)",
        "Notes": "Translates technical metrics into understandable environmental impact",
        "Suggested by": "Thomas"
    },
    {
        "Feature": "Generic Actionable Tips",
        "Short Description": "Display non-personalized sustainability tips that anyone can follow",
        "Prototype": "Magic Mirror (Prototype 1)",
        "Notes": "Avoids personalization for compliance while still providing value",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "Ministry-wide Leaderboard",
        "Short Description": "Friendly competition between departments on sustainability metrics",
        "Prototype": "Magic Mirror (Prototype 1)",
        "Notes": "Gamification element to drive engagement and healthy competition",
        "Suggested by": "Moses (based on feedback)"
    },
    {
        "Feature": "QR Codes for Resources",
        "Short Description": "Quick access to detailed information and guides via QR codes",
        "Prototype": "Magic Mirror (Prototype 1)",
        "Notes": "Bridges physical display with digital resources",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Run on Existing Screens",
        "Short Description": "Deploy on ministry's existing display infrastructure",
        "Prototype": "Magic Mirror (Prototype 1)",
        "Notes": "No new hardware required - reduces cost and complexity",
        "Suggested by": "Matthijs, Jop"
    },

    # Digital Forest Features
    {
        "Feature": "Clear Behavior Triggers",
        "Short Description": "Well-defined actions that lead to forest growth (tree planting)",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Addresses Thomas's key question: 'What behavior do you want to create?'",
        "Suggested by": "Thomas"
    },
    {
        "Feature": "Automated Tracking Architecture",
        "Short Description": "Backend system that automatically tracks sustainable actions without manual input",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Solves complexity concerns raised by all partners",
        "Suggested by": "Matthijs, Jop"
    },
    {
        "Feature": "Integration with Prompt Coach",
        "Short Description": "Connect forest growth to prompt optimization activities",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Creates unified experience across prototypes",
        "Suggested by": "Jop, Thomas"
    },
    {
        "Feature": "Integration with Dashboard",
        "Short Description": "Use dashboard analytics data to drive forest visualization",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Dashboard becomes 'analytical backbone' for visual prototypes",
        "Suggested by": "Thomas"
    },
    {
        "Feature": "Tiered Implementation (MVP→Automated→Real-time)",
        "Short Description": "Phased rollout starting simple and adding complexity over time",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Follows Matthijs's advice: start with minimal viable version",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "Behavior Matrix Documentation",
        "Short Description": "Clear mapping of which actions trigger which forest changes",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Addresses concerns about unclear behavior→outcome relationships",
        "Suggested by": "Thomas, Jop"
    },
    {
        "Feature": "Gamification Rules",
        "Short Description": "Point system, achievements, and rewards for sustainable actions",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Makes sustainability engaging and fun",
        "Suggested by": "Matthijs (gamification exercise)"
    },
    {
        "Feature": "Optimized Prompts Grow Trees",
        "Short Description": "Using optimized prompts from library triggers tree growth",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Links directly to Prompt Coach actions",
        "Suggested by": "Moses (based on feedback)"
    },
    {
        "Feature": "Quiz Completion Rewards",
        "Short Description": "Completing sustainability quizzes (>80%) grows trees",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Educational element with visual reward",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Efficiency Goal Achievement",
        "Short Description": "Achieving <1.5 prompt efficiency triggers forest growth",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Ties visual reward to concrete efficiency metric",
        "Suggested by": "Matthijs (efficiency focus)"
    },
    {
        "Feature": "Sustainable Alternative Adoption",
        "Short Description": "Switching to eco-friendly AI tools grows trees",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Encourages exploration of greener alternatives",
        "Suggested by": "Moses"
    },
    {
        "Feature": "CO₂ Reduction Milestones",
        "Short Description": "10% month-over-month CO₂ reduction triggers rewards",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Focuses on measurable environmental impact",
        "Suggested by": "Thomas (impact focus)"
    },
    {
        "Feature": "Prompt Sharing Rewards",
        "Short Description": "Sharing optimized prompts with colleagues grows trees",
        "Prototype": "Digital Forest (Prototype 2)",
        "Notes": "Encourages collaborative sustainability culture",
        "Suggested by": "Moses"
    },

    # Black Frame - Tetris Features
    {
        "Feature": "Tetris-style Grid Visualization",
        "Short Description": "Visual representation of AI usage intensity as filling Tetris blocks",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Jop's brilliant suggestion to replace manual drawing",
        "Suggested by": "Jop"
    },
    {
        "Feature": "Fully Automated Usage Tracking",
        "Short Description": "Automatic detection of AI requests without manual input",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Solves manual tracking problem that all partners flagged",
        "Suggested by": "Thomas, Jop"
    },
    {
        "Feature": "Every 5 AI Requests = +1 Block",
        "Short Description": "Clear ratio for how usage translates to visual blocks",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Makes abstract usage tangible and visible",
        "Suggested by": "Moses (implementing Jop's idea)"
    },
    {
        "Feature": "Color-coded Intensity Blocks",
        "Short Description": "Blocks change color based on usage intensity (low/moderate/high/critical)",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Visual feedback on usage severity",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Sustainable Actions Clear Lines",
        "Short Description": "Completing sustainable actions removes blocks like Tetris line clears",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Moses's addition to Jop's Tetris idea - makes it actionable",
        "Suggested by": "Moses (building on Jop's suggestion)"
    },
    {
        "Feature": "Gamified Clearing Rewards",
        "Short Description": "Point system and achievements for clearing blocks through sustainable actions",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Addresses Matthijs's actionability concern",
        "Suggested by": "Matthijs (actionability)"
    },
    {
        "Feature": "Real-time Activity Feed",
        "Short Description": "Live updates showing recent AI requests and clearing actions",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Transparency into what's driving the visualization",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Danger Zone Warning at 80%",
        "Short Description": "Alert when Tetris grid reaches critical capacity",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Creates urgency to take sustainable actions",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Optimize 5 Prompts = Clear 1 Line",
        "Short Description": "Specific clearing action: prompt optimization",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Links to Prompt Coach functionality",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Complete Quiz = Clear 2 Lines",
        "Short Description": "Sustainability quiz completion clears multiple lines",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Rewards educational engagement",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Use Eco Alternative = Clear 3 Lines",
        "Short Description": "Switching to sustainable AI tool clears significant blocks",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Higher reward for higher-impact action",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Achieve Efficiency Goal = Clear 5 Lines",
        "Short Description": "Meeting efficiency targets clears major portion of blocks",
        "Prototype": "Black Frame (Prototype 3)",
        "Notes": "Biggest reward for most impactful behavior change",
        "Suggested by": "Moses"
    },

    # Prompt Coach + Dashboard Features
    {
        "Feature": "Real-time CO₂ Tracking",
        "Short Description": "Show exact carbon emissions for each prompt (e.g., '0.32g CO₂')",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Thomas specifically mentioned wanting to see CO₂ usage",
        "Suggested by": "Thomas"
    },
    {
        "Feature": "AI-powered Optimization Suggestions",
        "Short Description": "Intelligent recommendations to improve prompt efficiency",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Core value proposition - helps users learn better prompts",
        "Suggested by": "Thomas"
    },
    {
        "Feature": "Quality vs. Efficiency Trade-off Indicator",
        "Short Description": "Visual display showing balance between result quality and computational efficiency",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Directly addresses Matthijs's key question about trade-offs",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "Three Priority Modes",
        "Short Description": "User-selectable modes: Quality First, Balanced, Efficiency First",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Gives users control over quality-efficiency balance",
        "Suggested by": "Moses (addressing Matthijs's concern)"
    },
    {
        "Feature": "Before/After Comparison",
        "Short Description": "Side-by-side view of original vs. optimized prompt with metrics",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Shows tangible improvement from optimization",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Savings Calculation",
        "Short Description": "Calculate CO₂, energy, and cost savings from optimization",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Quantifies impact of behavior change",
        "Suggested by": "Thomas (impact focus)"
    },
    {
        "Feature": "Environmental Impact Equivalents",
        "Short Description": "Translate CO₂ savings into relatable comparisons (e.g., 'X trees planted')",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Makes abstract numbers meaningful",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Reusable Prompt Library",
        "Short Description": "Collection of pre-optimized templates organized by category",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Jop made distinction between Coach and Library - wants both",
        "Suggested by": "Jop"
    },
    {
        "Feature": "Category Filtering",
        "Short Description": "Filter prompts by type: Code, Analysis, Writing, Research, etc.",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Makes library easy to navigate",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Prompt Search Functionality",
        "Short Description": "Search library by keywords, tags, or use case",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Quick access to relevant templates",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Usage Statistics Display",
        "Short Description": "Show how many others use each template",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Social proof encourages adoption of best practices",
        "Suggested by": "Moses"
    },
    {
        "Feature": "One-click Template Loading",
        "Short Description": "Instantly load template into prompt editor",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Jop said tool 'would save quite some time' - this enables that",
        "Suggested by": "Jop (time-saving)"
    },
    {
        "Feature": "Save Optimized Prompts",
        "Short Description": "Add newly optimized prompts to shared library",
        "Prototype": "Prompt Coach (Prototype 4)",
        "Notes": "Builds collective knowledge base",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Quick Stats Overview",
        "Short Description": "Dashboard widget showing CO₂ saved, efficiency score, queries optimized",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Jop wanted GAIA-style overview next to Prompt Coach",
        "Suggested by": "Jop"
    },
    {
        "Feature": "Goals & Progress Tracking",
        "Short Description": "Set sustainability targets and track progress over time",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Thomas emphasized defining goals and behaviors",
        "Suggested by": "Thomas"
    },
    {
        "Feature": "Achievements System",
        "Short Description": "Gamified rewards for sustainability milestones",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Engagement mechanism for behavior change",
        "Suggested by": "Matthijs (gamification)"
    },
    {
        "Feature": "Sustainable Alternatives Recommendations",
        "Short Description": "Suggest eco-friendly AI tool alternatives based on usage patterns",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Actionable suggestion to reduce impact",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Azure Metrics Integration",
        "Short Description": "Connect to Azure monitoring for real usage data",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Matthijs specifically said to use Azure data as starting point",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "Total Interactions Tracking",
        "Short Description": "Count all AI system interactions from Azure",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Foundation metric from existing infrastructure",
        "Suggested by": "Matthijs (use Azure)"
    },
    {
        "Feature": "Database Calls Monitoring",
        "Short Description": "Track database queries triggered by AI operations",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Infrastructure usage metric",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "Prompt Request Counting",
        "Short Description": "Total number of prompts sent to AI systems",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Core usage metric",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "CO₂ Estimations from Azure",
        "Short Description": "Leverage Azure's built-in carbon estimates",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Don't reinvent - use existing Azure capabilities",
        "Suggested by": "Matthijs, Thomas"
    },
    {
        "Feature": "Usage Pattern Heatmap",
        "Short Description": "Visualize when AI usage is highest, suggest time-shifting to greener hours",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Actionable insight for load balancing",
        "Suggested by": "Moses"
    },
    {
        "Feature": "Tool Diversity Tracking",
        "Short Description": "Monitor variety of AI tools being used",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Encourages exploring efficient alternatives",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "Efficiency Trends Over Time",
        "Short Description": "Chart showing improvement in prompt efficiency over weeks/months",
        "Prototype": "Dashboard (Prototype 5)",
        "Notes": "Shows long-term behavior change impact",
        "Suggested by": "Moses"
    },

    # Cross-cutting Features (Integration)
    {
        "Feature": "MVP-First Approach",
        "Short Description": "Start with minimal viable version, iterate based on usage",
        "Prototype": "All Prototypes",
        "Notes": "Matthijs's key advice: 'What is the bare minimum?'",
        "Suggested by": "Matthijs"
    },
    {
        "Feature": "Common Analytical Backbone",
        "Short Description": "Shared data layer across all visual prototypes",
        "Prototype": "Dashboard + All Visual Prototypes",
        "Notes": "Thomas said dashboard is 'the basis for all these things'",
        "Suggested by": "Thomas"
    },
    {
        "Feature": "Fewer Metrics, More Visualization",
        "Short Description": "Focus on visual impact rather than overwhelming data tables",
        "Prototype": "All Prototypes",
        "Notes": "Thomas's guidance on making dashboard engaging",
        "Suggested by": "Thomas"
    },
    {
        "Feature": "Link Visual Ideas with Data",
        "Short Description": "Connect creative visuals (forest, mirror) with analytical foundation",
        "Prototype": "Forest + Mirror + Dashboard",
        "Notes": "Thomas suggested combining visual appeal with data backbone",
        "Suggested by": "Thomas"
    },
    {
        "Feature": "Department-level Aggregation",
        "Short Description": "All personal data aggregated to department level for privacy compliance",
        "Prototype": "All Prototypes",
        "Notes": "Solves privacy/compliance concerns across all tools",
        "Suggested by": "Matthijs, Jop"
    },
]

# Create DataFrame
df = pd.DataFrame(features)

# Add row numbers
df.insert(0, '#', range(1, len(df) + 1))

# Reorder columns to match requested format
df = df[['#', 'Feature', 'Short Description', 'Prototype', 'Notes', 'Suggested by']]

# Create Excel file with formatting
excel_filename = 'PARTNER_FEEDBACK_FEATURES.xlsx'
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Features', index=False)

    # Get workbook and worksheet
    workbook = writer.book
    worksheet = writer.sheets['Features']

    # Set column widths
    worksheet.column_dimensions['A'].width = 5   # #
    worksheet.column_dimensions['B'].width = 35  # Feature
    worksheet.column_dimensions['C'].width = 60  # Short Description
    worksheet.column_dimensions['D'].width = 30  # Prototype
    worksheet.column_dimensions['E'].width = 60  # Notes
    worksheet.column_dimensions['F'].width = 25  # Suggested by

    # Format header row
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Format data cells
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Set row height for header
    worksheet.row_dimensions[1].height = 30

print(f"✅ Excel file created: {excel_filename}")
print(f"📊 Total features documented: {len(df)}")

# Also create a markdown version
markdown_filename = 'PARTNER_FEEDBACK_FEATURES.md'
with open(markdown_filename, 'w') as f:
    f.write("# Partner Feedback Features Table\n\n")
    f.write(f"**Generated**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
    f.write(f"**Total Features**: {len(df)}\n\n")
    f.write("## Summary by Prototype\n\n")

    # Count features by prototype
    prototype_counts = df['Prototype'].value_counts()
    for prototype, count in prototype_counts.items():
        f.write(f"- **{prototype}**: {count} features\n")

    f.write("\n---\n\n")
    f.write("## Complete Feature Table\n\n")

    # Write markdown table
    f.write(df.to_markdown(index=False))

    f.write("\n\n---\n\n")
    f.write("## Features by Prototype Category\n\n")

    # Group by prototype
    for prototype in df['Prototype'].unique():
        f.write(f"### {prototype}\n\n")
        prototype_df = df[df['Prototype'] == prototype][['#', 'Feature', 'Short Description', 'Suggested by']]
        f.write(prototype_df.to_markdown(index=False))
        f.write("\n\n")

    f.write("---\n\n")
    f.write("*This document was auto-generated from partner feedback analysis.*\n")
    f.write("*Source: PARTNER_FEEDBACK_SUMMARY.md*\n")

print(f"✅ Markdown file created: {markdown_filename}")

# Print summary
print("\n📋 Feature Breakdown:")
for prototype, count in prototype_counts.items():
    print(f"   {prototype}: {count} features")

print("\n✨ Files created successfully!")
print(f"   - {excel_filename}")
print(f"   - {markdown_filename}")
